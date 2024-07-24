import openpyxl
from openpyxl import load_workbook
from datetime import datetime

def budget_berechnung(einkommen):
    miete = 384  # Standardbetrag für Miete
    verträge = 100  # Standardbetrag für Verträge
    restbetrag = einkommen - miete - verträge
    essen_fitness = 250  # Standardbetrag für Essen und Fitness
    restbetrag -= essen_fitness
    bücher_weiterbildung = restbetrag * 0.10
    restbetrag -= bücher_weiterbildung
    urlaub = restbetrag * 0.10
    restbetrag -= urlaub
    familie = restbetrag * 0.10
    restbetrag -= familie
    sparen = restbetrag
    
    return {
        "Miete": miete,
        "Verträge": verträge,
        "Essen & Fitness": essen_fitness,
        "Bücher & Weiterbildung": bücher_weiterbildung,
        "Urlaub": urlaub,
        "Familie": familie,
        "Sparen": sparen
    }

def in_excel_speichern(daten, dateiname, monat):
    try:
        workbook = load_workbook(filename=dateiname)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Monatliches Budget"
        headers = ["Kategorie", "Miete", "Verträge", "Essen & Fitness", "Bücher & Weiterbildung", "Urlaub", "Familie", "Sparen", "Monat", "Datum"]
        sheet.append(headers)

    datum = datetime.now().strftime("%Y-%m-%d")
    
    row = [monat, datum] + [daten[kategorie] for kategorie in ["Miete", "Verträge", "Essen & Fitness", "Bücher & Weiterbildung", "Urlaub", "Familie", "Sparen"]]

    sheet.append(row)
    workbook.save(filename=dateiname)

def haupt():
    monat = input("Geben Sie den Namen des Monats ein: ")
    einkommen = float(input("Geben Sie Ihr monatliches Einkommen ein: "))
    
    budget = budget_berechnung(einkommen)
    
    print("\nBudgetübersicht:")
    for kategorie, betrag in budget.items():
        print(f"{kategorie}: {betrag:.2f}€")
    
    in_excel_speichern(budget, "Monatliches_Budget.xlsx", monat)
    print("\nBudget in 'Monatliches_Budget.xlsx' gespeichert")

if __name__ == "__main__":
    haupt()
